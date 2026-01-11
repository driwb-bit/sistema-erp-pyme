import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
# Importamos el modelo de ventas (OJO: recuerda la corrección de importación que hicimos antes)
from ventas.models import Venta 

@login_required
def reporte_ventas_excel(request):
    """
    Genera un Excel con todas las ventas del sistema y lo descarga automáticamente.
    """
    # 1. Crear el libro de trabajo (Workbook) y la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # 2. Definir los Encabezados de la tabla
    headers = ["ID Venta", "Fecha", "Cliente", "Vendedor", "Total ($)"]
    ws.append(headers)

    # 3. Dar estilo a la cabecera (Fondo Gris, Letra Negrita, Centrado)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Iteramos sobre la primera fila (la de los encabezados) para pintarla
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 4. Consultar la Base de Datos
    # Aquí podrías agregar filtros: Venta.objects.filter(fecha__month=12)
    ventas = Venta.objects.all().select_related('usuario').order_by('-fecha')

    # 5. Escribir los datos fila por fila
    for venta in ventas:
        # Formateamos la fecha para que no salga con milisegundos feos
        fecha_str = venta.fecha.strftime('%d/%m/%Y %H:%M')
        usuario_str = venta.usuario.username if venta.usuario else "Sistema"
        
        ws.append([
            venta.id,
            fecha_str,
            venta.cliente,
            usuario_str,
            float(venta.total) # Convertimos a float para que Excel lo sume bien
        ])

    # 6. Ajustar ancho de columnas automáticamente (estética)
    column_widths = [10, 20, 30, 20, 15]
    for i, column_width in enumerate(column_widths, 1): # 1-based index para columnas Excel
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # 7. Preparar la respuesta HTTP para descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Ventas_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

    # 8. Guardar el libro directamente en la respuesta HTTP
    wb.save(response)
    
    return response